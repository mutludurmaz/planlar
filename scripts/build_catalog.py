#!/usr/bin/env python3
"""planlar deposundaki .xlsx dosyalarindan catalog.json uretir."""

from __future__ import annotations

import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote

ROOT = Path(__file__).resolve().parents[1]
REPO = "mutludurmaz/planlar"
BRANCH = "main"
RAW_BASE = f"https://raw.githubusercontent.com/{REPO}/{BRANCH}"
OUT = ROOT / "catalog.json"
GRADE_RE = re.compile(r"(?<!\d)(9|10|11|12)(?!\d)")


def fold(text: str) -> str:
    n = unicodedata.normalize("NFKC", text).casefold()
    return (
        n.replace("ı", "i")
        .replace("i̇", "i")
        .replace("â", "a")
        .replace("ê", "e")
        .replace("î", "i")
        .replace("û", "u")
    )


def school_type(file_stem: str) -> tuple[str, str]:
    n = fold(file_stem)
    if "fen lise" in n:
        return "fen", "Fen Lisesi"
    if "anadolu lise" in n:
        return "anadolu", "Anadolu Lisesi"
    return "diger", file_stem.strip()


def posix_path(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def raw_url(rel: str) -> str:
    encoded = "/".join(quote(part) for part in rel.split("/"))
    return f"{RAW_BASE}/{encoded}"


def grades_from_text(text: str) -> list[int]:
    found: list[int] = []
    for match in GRADE_RE.finditer(text):
        grade = int(match.group(1))
        if grade not in found:
            found.append(grade)
    found.sort()
    return found


def grades_from_sheets(path: Path) -> list[int]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [9, 10, 11, 12]
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
    except Exception:
        return [9, 10, 11, 12]
    found: list[int] = []
    for name in names:
        n = fold(name)
        if "hafta" in n or "week" in n:
            continue
        match = GRADE_RE.search(name)
        if not match:
            continue
        grade = int(match.group(1))
        if grade not in found:
            found.append(grade)
    found.sort()
    return found or [9, 10, 11, 12]


def grades_for_file(path: Path) -> list[int]:
    from_name = grades_from_text(path.stem)
    if from_name:
        return from_name
    return grades_from_sheets(path)


def file_entry(path: Path, grades: list[int]) -> dict:
    rel = posix_path(path)
    return {
        "fileName": path.name,
        "path": rel,
        "downloadUrl": raw_url(rel),
        "grades": grades,
    }


def collapse_school_group(group: list[dict]) -> dict:
    if len(group) == 1:
        return group[0]
    ordered = sorted(
        group,
        key=lambda item: (min(item["grades"] or [99]), fold(item["fileName"])),
    )
    files = [
        {
            "grades": item["grades"],
            "fileName": item["fileName"],
            "path": item["path"],
            "downloadUrl": item["downloadUrl"],
        }
        for item in ordered
    ]
    grades: list[int] = []
    for item in ordered:
        for grade in item["grades"]:
            if grade not in grades:
                grades.append(grade)
    grades.sort()
    primary = max(
        ordered,
        key=lambda item: (len(item["grades"]), -min(item["grades"] or [99])),
    )
    return {
        "id": group[0]["id"],
        "title": group[0]["title"],
        "fileName": primary["fileName"],
        "path": primary["path"],
        "downloadUrl": primary["downloadUrl"],
        "grades": grades,
        "files": files,
    }


def merge_schools(schools: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = {}
    other: list[dict] = []
    for school in schools:
        if school["id"] in ("anadolu", "fen"):
            grouped.setdefault(school["id"], []).append(school)
        else:
            other.append(school)
    merged: list[dict] = []
    for school_id in ("anadolu", "fen"):
        group = grouped.get(school_id)
        if group:
            merged.append(collapse_school_group(group))
    merged.extend(other)
    ids = {school["id"] for school in merged}
    if "anadolu" in ids and "fen" not in ids:
        anadolu = next(school for school in merged if school["id"] == "anadolu")
        fen = json.loads(json.dumps(anadolu))
        fen["id"] = "fen"
        fen["title"] = "Fen Lisesi"
        merged.append(fen)
    return merged


def build() -> dict:
    by_subject: dict[str, dict] = {}
    files = sorted(
        (p for p in ROOT.rglob("*.xlsx") if ".git" not in p.parts),
        key=lambda p: (fold(p.parent.name), fold(p.name)),
    )
    for path in files:
        if path.parent == ROOT:
            continue
        subject = path.parent.name
        school_id, school_label = school_type(path.stem)
        grades = grades_for_file(path)
        school = {
            "id": school_id,
            "title": school_label,
            **file_entry(path, grades),
        }
        bucket = by_subject.setdefault(
            subject,
            {
                "title": subject,
                "path": path.parent.relative_to(ROOT).as_posix(),
                "schools": [],
            },
        )
        bucket["schools"].append(school)

    subjects = list(by_subject.values())
    subjects.sort(key=lambda item: fold(item["title"]))
    for subject in subjects:
        subject["schools"] = merge_schools(subject["schools"])
        subject["schools"].sort(
            key=lambda item: (
                0 if item["id"] == "anadolu" else 1 if item["id"] == "fen" else 2,
                fold(item["title"]),
            )
        )
    return {
        "version": 1,
        "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": f"https://github.com/{REPO}",
        "subjects": subjects,
    }


def main() -> None:
    catalog = build()
    OUT.write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    schools = sum(len(s["schools"]) for s in catalog["subjects"])
    files = sum(
        len(school.get("files") or [school])
        for subject in catalog["subjects"]
        for school in subject["schools"]
    )
    print(f"catalog.json: {len(catalog['subjects'])} ders, {schools} okul, {files} dosya")


if __name__ == "__main__":
    main()
